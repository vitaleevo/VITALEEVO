"""Endpoints de importação em massa — staff com catalog:import."""
from django.utils.text import slugify
from rest_framework import serializers, status, views
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response

from apps.audit.helpers import log_audit
from apps.catalog.models import Brand, Category, Product, ProductStatus
from apps.core.permissions import HasCapability

DEFAULT_PRODUCT_IMAGE = "/hero-card.png"

HEADER_ALIASES = {
    "sku": "sku",
    "name": "name",
    "nome": "name",
    "description": "description",
    "descricao": "description",
    "descrição": "description",
    "price": "price",
    "preco": "price",
    "preço": "price",
    "old_price": "old_price",
    "preco_antigo": "old_price",
    "preço_antigo": "old_price",
    "stock": "stock",
    "category": "category",
    "categoria": "category",
    "subcategory": "subcategory",
    "subcategoria": "subcategory",
    "brand": "brand",
    "marca": "brand",
    "image": "image",
    "imagem": "image",
    "slug": "slug",
    "is_featured": "is_featured",
    "destaque": "is_featured",
    "status": "status",
    "estado": "status",
}


class ImportProductsSerializer(serializers.Serializer):
    file = serializers.FileField()

    def validate_file(self, value):
        if not (value.name.endswith(".xlsx") or value.name.endswith(".xlsm")):
            raise serializers.ValidationError("Formato não suportado — use .xlsx.")
        if value.size > 10 * 1024 * 1024:
            raise serializers.ValidationError("Ficheiro excede 10 MB.")
        return value


def _parse_decimal(raw):
    if raw is None or str(raw).strip() == "":
        return None
    return float(str(raw).replace(",", ".").strip())


def _unique_category_slug(name: str, parent: Category | None) -> str:
    slug = slugify(name) or "categoria"
    candidate, counter = slug, 2
    while Category.objects.filter(slug=candidate).exists():
        candidate = f"{slug}-{counter}"
        counter += 1
    return candidate


def _resolve_category(name: str, parent: Category | None = None) -> Category:
    category = Category.objects.filter(name=name.strip(), parent=parent).first()
    if category:
        return category
    return Category.objects.create(
        name=name.strip(),
        slug=_unique_category_slug(name.strip(), parent),
        parent=parent,
        type=parent.type if parent else "store",
    )


def _resolve_brand(name: str) -> Brand:
    brand = Brand.objects.filter(name=name.strip()).first()
    if brand:
        return brand
    return Brand.objects.create(name=name.strip(), slug=_unique_slug(name.strip()))


def _unique_slug(base: str) -> str:
    slug = slugify(base) or "produto"
    candidate, counter = slug, 2
    while Product.objects.filter(slug=candidate).exists():
        candidate = f"{slug}-{counter}"
        counter += 1
    return candidate


class ProductImportView(views.APIView):
    """POST /imports/products/ — ficheiro Excel com produtos (cria/atualiza por SKU)."""

    parser_classes = [MultiPartParser, FormParser]

    def get_permissions(self):
        return [HasCapability("catalog:import")]

    def post(self, request):
        from openpyxl import load_workbook

        serializer = ImportProductsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            workbook = load_workbook(serializer.validated_data["file"], read_only=True, data_only=True)
        except Exception:  # noqa: BLE001
            return Response({"detail": "Não foi possível ler o ficheiro."}, status=status.HTTP_400_BAD_REQUEST)

        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Ficheiro vazio."}, status=status.HTTP_400_BAD_REQUEST)

        headers = {str(cell).strip().lower(): index for index, cell in enumerate(rows[0]) if cell is not None}
        mapped = {alias: headers[header] for header, alias in HEADER_ALIASES.items() if header in headers}

        created, updated, errors = 0, 0, []
        for row_index, raw in enumerate(rows[1:], start=2):
            if not any(raw):
                continue
            try:
                name = str(raw[mapped["name"]]).strip() if "name" in mapped else ""
                if not name:
                    raise ValueError("nome em falta")
                sku = str(raw[mapped["sku"]]).strip() if "sku" in mapped else ""
                price = _parse_decimal(raw[mapped["price"]]) if "price" in mapped else None
                if price is None or price < 0:
                    raise ValueError("preço inválido")

                category_name = str(raw[mapped["category"]]).strip() if "category" in mapped else "Geral"
                category = _resolve_category(category_name)
                subcategory = None
                if "subcategory" in mapped and str(raw[mapped["subcategory"]]).strip():
                    subcategory = _resolve_category(str(raw[mapped["subcategory"]]).strip(), parent=category)

                brand = None
                if "brand" in mapped and str(raw[mapped["brand"]]).strip():
                    brand = _resolve_brand(str(raw[mapped["brand"]]).strip())

                stock = int(float(raw[mapped["stock"]] or 0)) if "stock" in mapped and raw[mapped["stock"]] not in (None, "") else 0
                is_featured = bool(raw[mapped["is_featured"]]) if "is_featured" in mapped else False
                status_value = str(raw[mapped["status"]]).strip() if "status" in mapped else ProductStatus.PUBLISHED
                status_value = status_value if status_value in ProductStatus.values else ProductStatus.PUBLISHED
                image = str(raw[mapped["image"]]).strip() if "image" in mapped and raw[mapped["image"]] else DEFAULT_PRODUCT_IMAGE
                description = str(raw[mapped["description"]] or "").strip() if "description" in mapped else ""

                product = Product.objects.filter(sku=sku).first() if sku else None
                if product:
                    product.name = name
                    product.description = description
                    product.price = price
                    product.old_price = _parse_decimal(raw[mapped["old_price"]]) if "old_price" in mapped else product.old_price
                    product.stock = stock
                    product.image = image
                    product.category = category
                    product.subcategory = subcategory
                    product.brand = brand
                    product.is_featured = is_featured
                    product.status = status_value
                    product.save()
                    updated += 1
                else:
                    Product.objects.create(
                        name=name,
                        slug=_unique_slug(name),
                        sku=sku,
                        description=description,
                        price=price,
                        old_price=_parse_decimal(raw[mapped["old_price"]]) if "old_price" in mapped else None,
                        stock=stock,
                        image=image,
                        category=category,
                        subcategory=subcategory,
                        brand=brand,
                        is_featured=is_featured,
                        status=status_value,
                    )
                    created += 1
            except Exception as exc:  # noqa: BLE001 — falha de linha não aborta o lote
                errors.append({"row": row_index, "error": str(exc)})

        log_audit(
            user=request.user,
            action="catalog.import",
            resource_type="product",
            details={"created": created, "updated": updated, "errors": len(errors)},
        )
        return Response({"created": created, "updated": updated, "errors": errors})